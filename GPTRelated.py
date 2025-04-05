import re
import time
import openai
import pandas
from openpyxl import Workbook

# 设置API密钥
openai.api_key = "------------------hide-------------"

sentences = []


def get_metaphor_prompt(sentence):
    return f"""
    我想让你来做一些试题。我会提供一些句子，请你按照下面的标准判断每句话里是否存在隐喻，仔细按照以下五点打分的标准做出选择。
    句子：“{sentence}”
    请根据以下标准打分：
    （5）“这很明显是一个隐喻。”
    （4）“这应该是一个隐喻。”
    （3）“我不确定这是否是一个隐喻。”
    （2）“这可能是一种修辞用法。”
    （1）“这是语言的常见用法。”
    
    回答格式只需要：评分（无需重复题目）；原因（50字以内）
    """


def get_intention_prompt(sentence):
    return f"""
    接下来，你需要完成意图判断任务。
    针对每一个表达，你需要判断这一表达背后说话人可能的意图。请你从下面的意图分类中选择类别，每个隐喻仅能选择一项意图。您应该参考以下类别列表：
    （0）词汇化隐喻、（1）艺术隐喻、（2）可视化、（3）说服性、（4）解释性、（5）论证性隐喻、（6）社会互动、（7）幽默、（8）启发式推理、（9）特定文化内涵。
    
    句子：“{sentence}”
    请从列表中选择最合适的类别。

    回答格式只需要：序号；原因（50字以内）：xxxxxxx
    """


def get_V2metaphor_prompt(sentence):
    return f"""
    我想让你来做一些试题。我会提供一些句子，请你按照下面的标准判断每句话里是否存在隐喻，仔细按照以下五点打分的标准做出选择。
    句子：“{sentence}”
    请根据以下标准打分：
    （5）“这很明显是一个隐喻，而非句子的字面含义，源域与目标域的关系清晰。”
    （4）“这应该是一个隐喻，但有点缺乏说服力，源域与目标域的关系有一点不明显。”
    （3）“我不确定这是否是一个隐喻，源域与目标域之间可能存在关系。”
    （2）“这可能是一种隐喻用法，但我无法分辨。”
    （1）“这是语言的常见用法/完全无法发现源域与目标域之间的关系”
    
    回答格式只需要：评分（无需重复题目）；原因（50字以内）
    """


def get_V2intention_prompt(sentence):
    return f"""
    接下来，你需要完成意图判断任务。
    针对每一个表达，你需要判断这一表达背后说话人可能的意图。请你从下面的意图分类中选择类别，每个隐喻仅能选择一项意图。您应该参考以下类别列表：
    （0）词汇化隐喻 （1）艺术隐喻、（2）可视化、（3）说服性、（4）解释性、（5）论证性隐喻、（6）社会互动、（7）幽默、（8）启发式推理、（9）特定文化内涵。
    
    意图判断尽量只填写一种，以下是分类概述：
        0.	词汇化隐喻（Lexicalized metaphor）
        对于词汇化隐喻，在理解时不会涉及到为什么用隐喻表达而不用字面表达的疑问，他们主要是用人们惯常思维的模式传递命题信息。
        
        1.	艺术性隐喻（Artistic use of metaphor）
        这些隐喻用于一次性地向目标域赋予一整套特征。 这些特征不一定事先被明确规定。最终，使用者的意图是激发接收者的创造性解释。
        
        2.	可视化（Visualization）
        发话者可能会借助源域更容易被视觉化的隐喻。 意图是帮助接收者形成目标域的直观表征。
   
        3.	说服性隐喻（Persuasiveness）
        通过使用隐喻来指代目标域，作者可以赋予其非中立的内涵，而这种内涵并没有基于明确的论据。 其意图是让受众接受发话者对目标域的立场。
        
        4.	解释性隐喻（Explanation）
        这些隐喻用于教学目的。意图是向接收者解释一个新的或已经熟悉的概念。 在这一过程中，话语通常涉及从专家到非专家的知识不对称，例如从老师到学生。
        
        5.	论证性隐喻（Argumentative metaphor）
        这些隐喻是明确论据的一部分，作者旨在通过隐喻说服受众接受某种论断。 其意图是提出有力的陈述，无论是作为论据的立场还是起点（前提）。

        6.	社交互动隐喻（Social interaction）
        这些隐喻侧重于人际关系、群体或文化惯例等。其意图是创建或加强生产者与接收者之间的某种纽带。
            
        7.	幽默隐喻（Humour）
        其意图是娱乐接收者，产生幽默感。隐喻语言被用来产生有趣的效果，且在字面释义中无法实现。
   
        8.	启发式推理（Heuristic reasoning）
        其意图是为理论、艺术作品等提供一个解释模型，通常是用于难以构建或理解的抽象领域。隐喻表达被用来根据听者对源域的先前知识，帮助其组织对目标域的概念化。话语通常局限在专家之间。

        9.	文化特异性隐喻（Culturally Specific Metaphor）
        指的是在某一种特定文化中具有独特意义，扎根于特定文化的象征观念、历史习俗或社会环境，只有熟悉该文化背景的人才能准确理解。

    
    句子：“{sentence}”
    请从列表中选择最合适的类别。

    回答格式只需要：序号；原因（50字以内）：xxxxxxx
    """


def get_V3metaphor_prompt(sentence):
    return f"""
    我想让你来做一些试题。我会提供一些句子，请你按照下面的标准判断每句话里是否存在隐喻，仔细按照以下五点打分的标准做出选择。
    句子：“{sentence}”
    请根据以下标准打分：
    （5）“这很明显是一个隐喻，而非句子的字面含义，源域与目标域的关系清晰。”
    （4）“这应该是一个隐喻，但有点缺乏说服力，源域与目标域的关系有一点不明显。”
    （3）“我不确定这是否是一个隐喻。”
    （2）“这可能是一种修辞用法，但我无法分辨。”
    （1）“这是语言的常见用法/完全无法发现源域与目标域之间的关系”
    
    例如：
        “他在房间里看书。”————1
        “臭气一个劲儿的钻进鼻子。”————2
        “他在职场上是一匹黑马。”————3
        “他的心是一片荒凉的沙漠。”————4
        “你的心是柳叶落到水面般的静默。”————5
    回答格式只需要：评分（无需重复题目）；原因（50字以内）    

    """


def get_V3intention_prompt(sentence):
    return f"""
     接下来，你需要完成意图判断任务。
    针对每一个表达，你需要判断这一表达背后说话人可能的意图。请你从下面的意图分类中选择类别，每个隐喻仅能选择一项意图。您应该参考以下类别列表：
    （0）词汇化隐喻 （1）艺术隐喻、（2）可视化、（3）说服性、（4）解释性、（5）论证性隐喻、（6）社会互动、（7）幽默、（8）启发式推理、（9）特定文化内涵。
        
    意图判断尽量只填写一种，以下是分类概述：
        0.	词汇化隐喻（Lexicalized metaphor）
        对于词汇化隐喻，在理解时不会涉及到为什么用隐喻表达而不用字面表达的疑问，他们主要是用人们惯常思维的模式传递命题信息。
        用 Cameron 的话说，这类隐喻表达方式是“表达本身的方式”。
        鲁克伟、韦汉（2007）指出，死喻也即词汇化隐喻，是指隐喻义与该词的原义已经失去联系或成为该词的常用意义的一部分，使该词汇化隐喻缺乏非隐喻的其他替换性表达。
        例如：加油、十字路口、吃醋、鼠标等隐喻表达因其使用频率和认知的可及性提高，人们需要付出的认知努力大大减少，建立相对独立的神经元网络并固化。
        例如：
            a.	我坠入爱河。 b.争论是战争。
            
        1.	艺术性隐喻（Artistic use of metaphor）
        这些隐喻用于一次性地向目标域赋予一整套特征。 这些特征不一定事先被明确规定。最终，使用者的意图是激发接收者的创造性解释。
                例子：
            (a) 它是东方，而朱丽叶就是太阳。
            (b) 费米在物理学中的披风落到了年轻的肩膀上。
            
        2.	可视化（Visualization）
        发话者可能会借助源域更容易被视觉化的隐喻。 意图是帮助接收者形成目标域的直观表征。
         例子：
        (a) 就像有一道明亮的光正在向外照耀。
        (b) 它会像溜溜球一样上下弹跳。
        隐喻经常利用一个具体/可视化的源域来映射到一个抽象的目标域。这一点在表达主观感受时尤为明显。
        隐喻表达常被认为比字面释义更生动。因此，隐喻能够促使形成更具洞察力的心理图像。生动的隐喻不仅仅用于描述目的，也可以用于表达更清晰的指
   
        3.	说服性隐喻（Persuasiveness）
        通过使用隐喻来指代目标域，作者可以赋予其非中立的内涵，而这种内涵并没有基于明确的论据。 其意图是让受众接受发话者对目标域的立场。
        例子：
        (a) 伊斯兰浪潮。
        (b) 这部由著名诗人创作的瘦弱而无力的首部小说。
        隐喻通常突出目标域的某些方面，同时隐藏其他方面。这种突出与隐藏的过程在接收者中产生了框架效应，从而使目标域被“源域的扭曲镜头”所观察到。
        
        4.	解释性隐喻（Explanation）
        这些隐喻用于教学目的。意图是向接收者解释一个新的或已经熟悉的概念。 在这一过程中，话语通常涉及从专家到非专家的知识不对称，例如从老师到学生。
        例子：
        (a) 大气层是包围地球的气体毯。
        (b) 当中子崩解，吐出一个电子，它就变成了质子。
        
        5.	论证性隐喻（Argumentative metaphor）
        这些隐喻是明确论据的一部分，作者旨在通过隐喻说服受众接受某种论断。 其意图是提出有力的陈述，无论是作为论据的立场还是起点（前提）。
         例子：
          如果是这样的话，那将是一场赌博，因为他在沙特阿拉伯唯一的一次国际亮相中失败了。
        如van Poppel（2021）等人指出，论证性隐喻可以被用来提出有力的陈述，无论是作为论据的立场还是起点（前提）。此外，这些隐喻还能积极推动论证的进程。
        
        6.	社交互动隐喻（Social interaction）
        这些隐喻侧重于人际关系、群体或文化惯例等。其意图是创建或加强生产者与接收者之间的某种纽带。
        例子： 昏昏欲睡的乔，奸诈的希拉里。
        隐喻可以通过多种方式拉近生产者与欣赏者之间的距离。在这种情况下，社交隐喻被用来将预期的接收者与公众隔离开来（Cohen, 1978），从而强化群体内外动态。
        
        7.	幽默隐喻（Humour）
        其意图是娱乐接收者，产生幽默感。隐喻语言被用来产生有趣的效果，且在字面释义中无法实现。
        例子： 
        (a) 我是靴子世界里的地毯。
        (b) 你走进的地方我会称为储物间，但他们称其为浴室。
        语言不仅仅用于交流。在语言的多种用途中，还有一种是为了娱乐他人，并在此过程中获得娱乐。
        Steen（2008, 2014）引用了典型的幽默隐喻案例：体育报纸的头条、笑话、谜语等等。
        
        8.	启发式推理（Heuristic reasoning）
        其意图是为理论、艺术作品等提供一个解释模型，通常是用于难以构建或理解的抽象领域。隐喻表达被用来根据听者对源域的先前知识，帮助其组织对目标域的概念化。话语通常局限在专家之间。
        例子：
        (a) 气体就像一组随机运动的台球。
        (b) 她的身体像一块画布，她的外貌就是艺术。
        隐喻的本质在于“把某物看作某物”，即从某种角度来解释事物。从认知角度看，我们将源域映射到目标域，以便更好地理解它。
        
        9.	文化特异性隐喻（Culturally Specific Metaphor）
        指的是在某一种特定文化中具有独特意义，扎根于特定文化的象征观念、历史习俗或社会环境，只有熟悉该文化背景的人才能准确理解。
        例子：
        树高千丈，叶落归根
        万般皆下品，唯有读书高
    
    句子：“{sentence}”
    请从列表中选择最合适的类别。

    回答格式只需要：序号；原因（50字以内）：xxxxxxx
    """

def call_gpt(prompt, model="gpt-4o-mini"):
    response = openai.ChatCompletion.create(
        model=model,
        messages=[
            {"role": "user", "content": prompt}  # 用户的输入
        ]
    )

    # 返回第一条聊天回复
    return response['choices'][0]['message']['content'].strip()


# # 示例调用
# results = call_gpt("What is the capital of France?")
# print(results)


def find_numbers_in_string(s):
    # 使用正则表达式查找所有的数字
    numbers = re.findall(r'\（?(\d+)\）?', s)
    return numbers


def findExplianString(s):
    match = re.search(r'\（?(\d+)\）?', s)
    if match is not None:
        start = match.start(1)
        # 从这个位置截取字符串直到末尾
        substring = s[start + 2:]
        return substring
    return ""


def excelOption(responseList, flag):
    if flag == "W":
        wb = Workbook()
        ws = wb.active

        # 添加数据到工作表
        ws['A1'] = 'intentionScore'
        ws['B1'] = 'metaphorScore'
        ws['C1'] = 'metaphorExplain'
        ws['D1'] = 'GPT_Int_Response'
        ws['E1'] = 'GPT_Meta_Response'

        for item in responseList:
            ws.append(item)

        # 保存工作簿
        wb.save('output1029.xlsx')
        return 0
    else:
        df = pandas.read_excel(
            "./MyData.xlsx",
            sheet_name='gpt测试语料',
            usecols=[0],
            engine='openpyxl'
        )
        return df.values.tolist()

def excelOptionV1(responseList, flag):
    if flag == "W":
        wb = Workbook()
        ws = wb.active

        # 添加数据到工作表
        ws['A1'] = 'Sentence'
        ws['B1'] = 'Score'
        ws['C1'] = 'GPT_Meta_Response'
        # ws['D1'] = 'GPT_Int_Response'
        # ws['E1'] = 'GPT_Meta_Response'

        for item in responseList:
            ws.append(item)

        # 保存工作簿
        wb.save('output-v1.xlsx')
        return 0
    else:
        df = pandas.read_excel(
            "./MyData.xlsx",
            sheet_name='Sheet1',
            usecols=[0],
            engine='openpyxl'
        )
        return df.values.tolist()

def excelOptionV2(responseList, flag):
    if flag == "W":
        wb = Workbook()
        ws = wb.active

        # 添加数据到工作表
        ws['A1'] = 'Sentence'
        ws['B1'] = 'Score'
        ws['C1'] = 'GPT_Meta_Response'
        # ws['D1'] = 'GPT_Int_Response'
        # ws['E1'] = 'GPT_Meta_Response'

        for item in responseList:
            ws.append(item)

        # 保存工作簿
        wb.save('output-v2.xlsx')
        return 0
    else:
        df = pandas.read_excel(
            "./MyData.xlsx",
            sheet_name='Sheet1',
            usecols=[0],
            engine='openpyxl'
        )
        return df.values.tolist()

def excelOptionV3(responseList, flag):
    if flag == "W":
        wb = Workbook()
        ws = wb.active

        # 添加数据到工作表
        ws['A1'] = 'Sentence'
        ws['B1'] = 'Score'
        ws['C1'] = 'GPT_Meta_Response'
        # ws['D1'] = 'GPT_Int_Response'
        # ws['E1'] = 'GPT_Meta_Response'

        for item in responseList:
            ws.append(item)

        # 保存工作簿
        wb.save('output-v3.xlsx')
        return 0
    else:
        df = pandas.read_excel(
            "./MyData.xlsx",
            sheet_name='gpt测试语料',
            usecols=[0],
            engine='openpyxl'
        )
        return df.values.tolist()

def getExcelDataV2( metaphor_response,stc):
    singleResponseFromGPT = (
        stc,
        find_numbers_in_string(metaphor_response)[0],
        metaphor_response
    )
    return singleResponseFromGPT


def getExcelData(intention_response, metaphor_response):
    singleResponseFromGPT = (
        find_numbers_in_string(intention_response)[0],
        findExplianString(metaphor_response),
        find_numbers_in_string(metaphor_response)[0],
        intention_response,
        metaphor_response
    )
    return singleResponseFromGPT


def getResponseBySentence(sentence, type):
    if type == "metaphor":
        prompt = get_metaphor_prompt(sentence)
    else:
        prompt = get_intention_prompt(sentence)

    response = call_gpt(prompt)
    return response
def getResponseBySentenceV2(sentence,type):
    if type=="V1":
        prompt = get_metaphor_prompt(sentence)
    if type=="V2":
        prompt = get_V2metaphor_prompt(sentence)
    if type=="V3":
        prompt = get_V3metaphor_prompt(sentence)

    response = call_gpt(prompt)
    return response

def getResponseBySentenceInt(sentence,type):
    if type=="V1":
        prompt = get_intention_prompt(sentence)
    if type=="V2":
        prompt = get_V2intention_prompt(sentence)
    if type=="V3":
        prompt = get_V3intention_prompt(sentence)

    response = call_gpt(prompt)
    return response
if __name__ == '__main__':

    # set method's mode as R(Read), resolve and load the sentences from 语料库excel
    sentences = excelOptionV2(None, "R")

    # set a variable to save the result in future
    responseList1=[]
    responseList2 = []
    responseList3=[]
    indeVal=0
    # Iterate all the sentences and get response from GPT
    for sentence in sentences:

        # add try catch to prevent any error occur
        try:
            indeVal=indeVal+1
            # get metaphor_response
            metaphor_response = getResponseBySentence(sentence, "metaphor")
            metaphor_response2 = getResponseBySentenceV2(sentence, "V2")
            metaphor_response3 = getResponseBySentenceV2(sentence, "V3")


            # get intention_response
            #intention_response = getResponseBySentence(sentence, "intention")

            # new_response=GetResponseBySentence(sentence,"new")

            # add log so that console panel can see
            print(f"\nA Item Done---{indeVal}-----", "\nmetaphor_response:",
                  int1Response1,int1Response2,int1Response3)

            # input intention and metaphor response,compose to single Excel row data
            singleRowData1 = getExcelDataV2(metaphor_response,sentence[0])
            singleRowData2 = getExcelDataV2(metaphor_response2,sentence[0])
            singleRowData3 = getExcelDataV2(metaphor_response3,sentence[0])
            # add data to a list
            responseList1.append(singleRowData2)
            responseList2.append(singleRowData2)
            responseList3.append(singleRowData3)

            # 避免频繁请求被API限流
            time.sleep(1)
        except Exception as e:

            # when error occur, save the data to the Excel
            excelOptionV1(responseList1, "W")
            excelOptionV2(responseList2, "W")
            excelOptionV3(responseList3, "W")

    # when execute finished, save result to Excel
    excelOptionV1(responseList1, "W")
    excelOptionV2(responseList2, "W")
    excelOptionV3(responseList3, "W")
